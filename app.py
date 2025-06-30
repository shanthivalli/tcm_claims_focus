import streamlit as st
import pandas as pd
from datetime import datetime, date, time as dt_time
import traceback
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import json
import hashlib
from typing import Dict
from pathlib import Path
from openpyxl import load_workbook
import time
from io import BytesIO
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import difflib
import pytz
from fpdf import FPDF
import base64

def entry_label(entry):
                        member_name = entry.get('member_name') or entry.get('MEMBER NAME') or 'N/A'
                        medicaid_id = entry.get('medicaid_id') or entry.get('MEDICAID ID') or 'N/A'
                        service_date = entry.get('service_date') or entry.get('DATE OF SERVICE') or 'N/A'
                        note_category = entry.get('note_category') or entry.get('Type of Note') or 'N/A'
                        return f"{member_name} - {medicaid_id} | {service_date} | {note_category}"

# Define all constants at the top
TOTAL_SECTIONS = 8  # Updated to include the new demographic section
ADMIN_OPTIONS = ["View Forms"]
ADMIN_CREDENTIALS = {
    "admin": hashlib.sha256("admin123".encode()).hexdigest(),
    "supervisor": hashlib.sha256("super456".encode()).hexdigest(),
    "w.turano@focuscares.com": hashlib.sha256("Focus@2024".encode()).hexdigest()
}
UNIFORM_TC_PASSWORD = "Focus2024!"  # Uniform password for all TCs

REQUIRED_FIELDS = [
    "Id", "Start time", "Completion time", "Email", "Name", "Is this a new note or amendment?", "REASON FOR AMENDMENT", "DATE OF SERVICE", "Did you travel to/for client?", "TOTAL CLIENT TRAVEL TIME", "CLIENT TRAVEL DETAILS", "Type of Note", "MEDICAID ID", "MEMBER NAME", "MEMBER ID", "MEMBER DOB", "TCM HOURS-ENTER", "TCM UNITS/Minutes", "ICD 10", "CPT CODE", "TOTAL TRAVEL TIME", "OUTLINE EACH DEST", "ADDITIONAL COMMENTS", "TRANSITION COORDINATION TASK COMPLETED", "NEXT STEPS/PLAN FOR", "TYPE OF CONTACT"
]

# Field mapping from form data to required fields
FORM_TO_REQUIRED_MAPPING = {
    # Basic demographics
    "medicaid_id": "MEDICAID ID",
    "member_name": "MEMBER NAME", 
    "member_id": "MEMBER ID",
    "member_dob": "MEMBER DOB",
    "tc_name": "Name",
    "tc_email": "Email",
    "service_date": "DATE OF SERVICE",
    "start_time": "Start time",
    "end_time": "Completion time",
    
    # Note information
    "note_type": "Is this a new note or amendment?",
    "amendment_reason": "REASON FOR AMENDMENT",
    "note_category": "Type of Note",
    
    # Travel information
    "travel_to_client": "Did you travel to/for client?",
    "travel_time": "TOTAL CLIENT TRAVEL TIME",
    "travel_details": "CLIENT TRAVEL DETAILS",
    "total_travel_time": "TOTAL TRAVEL TIME",
    "travel_locations": "OUTLINE EACH DEST",
    "travel_comments": "ADDITIONAL COMMENTS",
    
    # TCM specific
    "tcm_hours": "TCM HOURS-ENTER",
    "tcm_units": "TCM UNITS/Minutes",
    "icd_10": "ICD 10",
    "cpt_code": "CPT CODE",
    
    # Tasks and contacts
    "tasks_completed": "TRANSITION COORDINATION TASK COMPLETED",
    "next_steps": "NEXT STEPS/PLAN FOR",
    "contact_types": "TYPE OF CONTACT",
    
    # Administrative
    "admin_type": "Type of Note",
    "admin_comments": "ADDITIONAL COMMENTS",
    
    # Default values for fields not in form
    "Id": "Id"
}

def dict_to_pdf(data: dict, title="Form Details"):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    page_width = pdf.w - 2 * pdf.l_margin  # Calculate usable page width
    pdf.cell(page_width, 10, txt=title, ln=True, align='C')
    pdf.ln(10)
    for key, value in data.items():
        line = f"{key}: {value}"
        pdf.multi_cell(page_width, 10, line)
    return pdf.output(dest='S').encode('latin1')




def map_form_data_to_required_fields(form_data):
    """
    Map form data field names to required field names.
    Returns a dictionary with required field names as keys.
    """
    mapped_entry = {}
    
    # Map form data to required fields
    for form_field, required_field in FORM_TO_REQUIRED_MAPPING.items():
        if form_field in form_data:
            value = form_data[form_field]
            
            # Handle special data type conversions
            if required_field == "MEMBER ID" and value:
                try:
                    mapped_entry[required_field] = int(float(value)) if value else 0
                except (ValueError, TypeError):
                    mapped_entry[required_field] = 0
            elif required_field in ["Start time", "Completion time"] and value:
                # Convert time objects to string format
                if hasattr(value, 'strftime'):
                    mapped_entry[required_field] = value.strftime("%H:%M")
                else:
                    mapped_entry[required_field] = str(value)
            elif required_field == "DATE OF SERVICE" and value:
                # Convert date objects to string format
                if hasattr(value, 'strftime'):
                    mapped_entry[required_field] = value.strftime("%m/%d/%Y")
                else:
                    mapped_entry[required_field] = str(value)
            elif required_field == "MEMBER DOB" and value:
                # Convert date objects to string format
                if hasattr(value, 'strftime'):
                    mapped_entry[required_field] = value.strftime("%m/%d/%Y")
                else:
                    mapped_entry[required_field] = str(value)
            elif required_field == "TYPE OF CONTACT" and isinstance(value, list):
                # Convert list to string for contact types
                mapped_entry[required_field] = ", ".join(value) if value else "none"
            elif required_field == "ICD 10" and isinstance(value, bool):
                # Convert boolean to string
                mapped_entry[required_field] = "Yes" if value else "No"
            else:
                mapped_entry[required_field] = value
        else:
            # Set default values for missing fields (but not for time fields)
            if required_field in ["Id"]:
                mapped_entry[required_field] = "none"
            elif required_field in ["MEMBER ID"]:
                mapped_entry[required_field] = 0
            elif required_field == "TYPE OF CONTACT":
                mapped_entry[required_field] = "none"
            elif required_field == "ICD 10":
                mapped_entry[required_field] = "No"
            elif required_field in ["Start time", "Completion time"]:
                # Don't set default values for time fields - let them remain empty
                mapped_entry[required_field] = ""
            else:
                mapped_entry[required_field] = "none"
    
    # Ensure all required fields are present
    for field in REQUIRED_FIELDS:
        if field not in mapped_entry:
            if field in ["Id"]:
                mapped_entry[field] = "none"
            elif field in ["MEMBER ID"]:
                mapped_entry[field] = 0
            elif field == "TYPE OF CONTACT":
                mapped_entry[field] = "none"
            elif field == "ICD 10":
                mapped_entry[field] = "No"
            elif field in ["Start time", "Completion time"]:
                # Don't set default values for time fields - let them remain empty
                mapped_entry[field] = ""
            else:
                mapped_entry[field] = "none"
    
    return mapped_entry

# Helper to ensure all required fields are present in the entry
def fill_missing_fields(entry, required_fields, numeric_fields=None):
    if numeric_fields is None:
        numeric_fields = []
    for field in required_fields:
        if field not in entry or entry[field] in [None, '', [], {}]:
            entry[field] = 0 if field in numeric_fields else 'none'
    return entry

# Initialize ALL session state variables at the start of the script
if 'log_entries' not in st.session_state:
    try:
        with open('log_entries.json', 'r') as f:
            st.session_state.log_entries = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        st.session_state.log_entries = []

if 'current_section' not in st.session_state:
    st.session_state.current_section = 0

if 'form_data' not in st.session_state:
    st.session_state.form_data = {}

if 'member_data' not in st.session_state:
    st.session_state.member_data = {}

if 'member_verified' not in st.session_state:
    st.session_state.member_verified = False

if 'new_member' not in st.session_state:
    st.session_state.new_member = False

if 'is_admin' not in st.session_state:
    st.session_state.is_admin = False

if 'nav_selection' not in st.session_state:
    st.session_state.nav_selection = "Member Login"


if 'admin_selection' not in st.session_state:
    st.session_state.admin_selection = "View Forms"

# Add these to your other session state initializations
if 'service_date_checked' not in st.session_state:
    st.session_state.service_date_checked = False

if 'duplicate_service_date_confirmed' not in st.session_state:
    st.session_state.duplicate_service_date_confirmed = False

if 'selected_service_date' not in st.session_state:
    st.session_state.selected_service_date = None


def save_entries():
    """Save all log entries to a JSON file."""
    try:
        # Remove automatic time setting - let users manually enter times
        # Only save entries as they are, without modifying times
        with open('log_entries.json', 'w') as f:
            json.dump(st.session_state.log_entries, f, default=str, indent=4)
        return True
    except Exception as e:
        st.error(f"Error saving entries: {str(e)}")
        return False
      

def parse_time(val, default=dt_time(9, 0)):
    if isinstance(val, dt_time):
        return val
    if isinstance(val, str):
        try:
            return dt_time.strptime(val, "%H:%M").time()
        except Exception:
            pass
    return default
  

def validate_medicaid_id(medicaid_id: str) -> tuple[bool, str]:
    """
    Validate Medicaid ID format:
    - Must be 7 characters long
    - First character must be a letter
    - Can contain both letters and numbers
    Returns: (is_valid: bool, error_message: str)
    """
    if not medicaid_id:
        return False, "Medicaid ID is required"
    
    # Remove any spaces and special characters
    medicaid_id = medicaid_id.strip().upper()
    
    # Check length
    if len(medicaid_id) != 7:
        return False, "Medicaid ID must be exactly 7 characters long"
    
    # Check if first character is a letter
    if not medicaid_id[0].isalpha():
        return False, "Medicaid ID must start with a letter"
    
    # Check if remaining characters are alphanumeric
    if not medicaid_id[1:].isalnum():
        return False, "Medicaid ID can only contain letters and numbers"
    
    # If we reach here, the format is valid
    return True, ""


def get_member_details(medicaid_id: str) -> Dict:
    """
    Get member details from Excel file based on Medicaid ID.
    Returns a dictionary with member details or empty dict if not found.
    """
    try:
        # Load the Excel file - make sure to use the correct file path
        excel_path = './Master_db.xlsx'  # Update this to match your actual file name
        
        if not Path(excel_path).exists():
            st.error(f"Excel file not found at: {excel_path}")
            return {}
            
        # Read Excel file
        df = pd.read_excel(excel_path)
        
        # Convert MedicaidID to string and clean it
        df['MedicaidID'] = df['MedicaidID'].astype(str).str.strip()
        medicaid_id = str(medicaid_id).strip()
        
        # Find the member (case-insensitive comparison)
        member_mask = df['MedicaidID'].str.upper() == medicaid_id.upper()
        
        if not member_mask.any():
            return {}
            
        # Get the first matching row
        member_row = df[member_mask].iloc[0]
        
        # Create a dictionary with standardized keys
        member_dict = {
            'medicaid_id': medicaid_id,
            'member_name': member_row.get('FIRST NAME', '') + ' ' + member_row.get('LAST NAME', ''),
            'member_dob': member_row.get('DOB', None),
            'member_id': member_row.get('MEMBER ID', ''),
            'tc_name': member_row.get('TRANSITION COORDINATOR', ''),
            'tc_email': member_row.get('TC EMAIL', '')
        }
        
        return member_dict
    except Exception as e:
        st.error(f"Error retrieving member details: {str(e)}")
        print(f"Error retrieving member details: {str(e)}")  # Debug log
        return {}

# Set page configuration
st.set_page_config(
    page_title="2025 Colorado Transition Coordinator Log Notes",
    page_icon="📝",
    layout="wide",
    initial_sidebar_state="auto",  # can be "auto", "expanded", "collapsed"
    menu_items={
        'Get Help': None,
        'Report a bug': None,
        'About': None
    }
)

with open("custom_styles.css") as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)


with st.sidebar:
    st.image("https://focuscares.com/wp-content/uploads/elementor/thumbs/logo-pcsu71jmplrr1f3b7mtv083rbyula7p5imfik70y8o.png", width=160)
    # Add a separator line
    st.markdown("<hr style='margin-top: 0; margin-bottom: 1rem;'>", unsafe_allow_html=True)



    
    if st.button("Member Login", 
                type="primary" if st.session_state.nav_selection == "Member Login" else "secondary",
                use_container_width=True):
        st.session_state.nav_selection = "Member Login"
        st.session_state.admin_selection = None
        st.rerun()

    if st.button("Support", 
                type="primary" if st.session_state.nav_selection == "Support" else "secondary",
                use_container_width=True):
        st.session_state.nav_selection = "Support"
        st.session_state.admin_selection = None
        st.rerun()

    if st.button("Admin", 
                type="primary" if st.session_state.nav_selection == "Admin" else "secondary",
                use_container_width=True):
        st.session_state.nav_selection = "Admin"
        st.session_state.admin_selection = None
        st.rerun()


# Show admin sub-navigation if Admin is selected and user is admin
if st.session_state.nav_selection == "Admin" and st.session_state.is_admin:
    admin_row = st.container()
    with admin_row:
        admin_cols = st.columns(4)  # Change to 4 columns for 4 tabs
        
        with admin_cols[0]:
            if st.button("View Forms", 
                        key="view_forms", 
                        type="primary" if st.session_state.admin_selection == "View Forms" else "secondary",
                        use_container_width=True):
                st.session_state.admin_selection = "View Forms"
                st.rerun()
        
        
    # Add another separator line after admin sub-navigation if needed
if st.session_state.nav_selection == "Admin" and st.session_state.is_admin:
    st.markdown("<hr style='margin-top: 0.5rem; margin-bottom: 1rem;'>", unsafe_allow_html=True)

# Add this code to handle the admin selections
if st.session_state.nav_selection == "Admin" and st.session_state.is_admin:
    # Clear member-specific session state
    st.session_state.member_verified = False
    st.session_state.member_data = {}
    st.session_state.form_data = {}
    st.session_state.current_section = 0
    st.session_state.service_date_checked = False
    st.session_state.duplicate_service_date_confirmed = False
    
    if st.session_state.get('admin_selection') == "View Forms":  # Changed from "View Submitted Forms"
        st.markdown('<h2 class="subheader">View Submitted Forms</h2>', unsafe_allow_html=True)
        
        # Load log entries
        try:
            with open('log_entries.json', 'r') as f:
                log_entries = json.load(f)
            
            if log_entries:
                # Create filters
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    # Get unique Medicaid IDs
                    medicaid_ids = sorted(list(set(entry.get('medicaid_id', '') for entry in log_entries if 'medicaid_id' in entry)))
                    selected_medicaid_id = st.selectbox("Filter by Medicaid ID", ["All"] + medicaid_ids)
                
                with col2:
                    # Get unique note categories
                    note_categories = sorted(list(set(entry.get('note_category', '') for entry in log_entries if 'note_category' in entry)))
                    selected_category = st.selectbox("Filter by Note Category", ["All"] + note_categories)
                
                with col3:
                    # Date range filter
                    date_range = st.date_input(
                        "Filter by Date Range",
                        value=(datetime.now().date() - pd.Timedelta(days=30), datetime.now().date()),
                        max_value=datetime.now().date(),
                        format="MM/DD/YYYY"
                    )
                
                # Apply filters
                filtered_entries = log_entries
                
                # Filter by Medicaid ID
                if selected_medicaid_id != "All":
                    filtered_entries = [
                        entry for entry in filtered_entries 
                        if (entry.get('medicaid_id') == selected_medicaid_id or 
                            entry.get('MEDICAID ID') == selected_medicaid_id)
                    ]
                
                # Filter by note category
                if selected_category != "All":
                    filtered_entries = [
                        entry for entry in filtered_entries 
                        if (entry.get('note_category') == selected_category or
                            entry.get('Type of Note') == selected_category)
                    ]
                
                # Filter by date range
                if len(date_range) == 2:
                    start_date, end_date = date_range
                    temp_filtered_entries = []
                    
                    for entry in filtered_entries:
                        # Try different field names for service date
                        service_date_str = (entry.get('service_date') or 
                                          entry.get('DATE OF SERVICE') or 
                                          '01/01/1900')
                        try:
                            # Try mm/dd/yyyy format first
                            service_date = datetime.strptime(service_date_str, '%m/%d/%Y').date()
                        except ValueError:
                            try:
                                # Try yyyy-mm-dd format as fallback
                                service_date = datetime.strptime(service_date_str, '%Y-%m-%d').date()
                            except ValueError:
                                # Skip entries with invalid dates
                                continue
                        
                        if start_date <= service_date <= end_date:
                            temp_filtered_entries.append(entry)
                    
                    filtered_entries = temp_filtered_entries
                
                # Create a DataFrame for display
                if filtered_entries:
                    # Use entries as they are without patching times
                    # Remove automatic time patching - let users manually enter times
                    patched_entries = filtered_entries.copy()

                    # Use the same desired order as the member table
                    desired_order = [
                        "Start time", "Completion time", "Email", "Name", "Is this a new note or amendment?", "REASON FOR AMENDMENT",
                        "DATE OF SERVICE", "Did you travel to/for client?", "TOTAL CLIENT TRAVEL TIME", "CLIENT TRAVEL DETAILS",
                        "TOTAL TRAVEL TIME", "OUTLINE EACH DEST", "ADDITIONAL COMMENTS", "TCM HOURS-ENTER", "TCM UNITS/Minutes",
                        "ICD 10", "CPT CODE", "TRANSITION COORDINATION TASK COMPLETED", "NEXT STEPS/PLAN FOR", "TYPE OF CONTACT",
                        "MEDICAID ID", "MEMBER NAME", "MEMBER ID", "MEMBER DOB"
                    ]
                    # Collect all unique keys from ALL entries in the JSON, not just filtered ones
                    all_keys = set()
                    for entry in log_entries:
                        all_keys.update(entry.keys())
                    # Add any extra columns at the end
                    ordered_cols = desired_order + [col for col in all_keys if col not in desired_order]

                    # Build table data
                    def get_best_value(entry, col):
                        # Try exact match
                        if col in entry and entry[col] not in [None, '', [], {}]:
                            return entry[col]
                        # Try lower/underscore version
                        alt_col = col.lower().replace(' ', '_')
                        if alt_col in entry and entry[alt_col] not in [None, '', [], {}]:
                            return entry[alt_col]
                        # Try upper/space version
                        alt_col2 = col.upper().replace('_', ' ')
                        if alt_col2 in entry and entry[alt_col2] not in [None, '', [], {}]:
                            return entry[alt_col2]
                        return ''

                    table_data = []
                    for entry in patched_entries:
                        row = {col: get_best_value(entry, col) for col in ordered_cols}
                        table_data.append(row)

                    # Replace empty string values with 'none' for display
                    for row in table_data:
                        for k, v in row.items():
                            if v == "":
                                row[k] = "none"

                    # Format MEMBER DOB as MM/DD/YYYY for display
                    for row in table_data:
                        dob = row.get("MEMBER DOB", "") or row.get("member_dob", "")
                        if dob and dob != "none":
                            try:
                                dob_dt = pd.to_datetime(dob)
                                row["MEMBER DOB"] = dob_dt.strftime("%m/%d/%Y")
                                row["member_dob"] = dob_dt.strftime("%m/%d/%Y")
                            except Exception:
                                pass

                    df = pd.DataFrame(table_data, columns=ordered_cols)
                    st.dataframe(df, use_container_width=True)
                    st.markdown("### View Form Details")

                    # Build a readable label for each entry
                    
                    selected_entry_idx = st.selectbox(
                        "Select a form to view details",
                        options=range(len(patched_entries)),
                        format_func=lambda i: entry_label(patched_entries[i])
                    )

                    # --- EDIT FEATURE START ---
                    if 'admin_edit_idx' not in st.session_state:
                        st.session_state.admin_edit_idx = None
                    if 'edit_form_data' not in st.session_state:
                        st.session_state.edit_form_data = None

                    colA, colB = st.columns([1, 1])
                    with colA:
                        if st.button("View Details", key="view_details_btn"):
                            selected_entry = patched_entries[selected_entry_idx]
                            pdf_bytes = dict_to_pdf(selected_entry, title="Form Details")
                            b64_pdf = base64.b64encode(pdf_bytes).decode('utf-8')
                            pdf_display = f'<iframe src="data:application/pdf;base64,{b64_pdf}" width="700" height="900" type="application/pdf"></iframe>'
                            st.markdown(pdf_display, unsafe_allow_html=True)
                            st.download_button(
                                label="Download PDF",
                                data=pdf_bytes,
                                file_name="form_details.pdf",
                                mime="application/pdf"
                            )
                    with colB:
                        if st.button("Edit", key=f"edit_btn_{selected_entry_idx}"):
                            st.session_state.admin_edit_idx = selected_entry_idx
                            st.session_state.edit_form_data = patched_entries[selected_entry_idx].copy()
                            st.rerun()

                    # Show edit form if an entry is selected for editing
                    if st.session_state.admin_edit_idx is not None:
                        edit_idx = st.session_state.admin_edit_idx
                        edit_data = st.session_state.get('edit_form_data', patched_entries[edit_idx].copy())
                        member_name = edit_data.get('MEMBER NAME', edit_data.get('member_name', ''))
                        medicaid_id = edit_data.get('MEDICAID ID', edit_data.get('medicaid_id', ''))
                        st.markdown(f"### Edit Form Entry for: <span style='color:#1976d2'><b>{member_name}</b></span> (<span style='color:#1976d2'><b>{medicaid_id}</b></span>)", unsafe_allow_html=True)
                        with st.form("admin_edit_form"):
                            updated_data = {}
                            for field in ordered_cols:
                                val = edit_data.get(field, "")
                                updated_data[field] = st.text_input(field, value=str(val) if val is not None else "", key=f"edit_{field}")
                            submitted = st.form_submit_button("Save Changes")
                            if submitted:
                                # Save in REQUIRED_FIELDS order, then extras
                                new_entry = {field: updated_data.get(field, "") for field in desired_order}
                                for field in ordered_cols:
                                    if field not in desired_order:
                                        new_entry[field] = updated_data.get(field, "")
                                st.session_state.log_entries[edit_idx] = new_entry
                                save_entries()
                                st.success("Entry updated successfully!")
                                st.session_state.admin_edit_idx = None
                                st.session_state.edit_form_data = None
                                st.rerun()
                        if st.button("Cancel Edit", key="cancel_edit_btn"):
                            st.session_state.admin_edit_idx = None
                            st.session_state.edit_form_data = None
                            st.rerun()
                    # --- EDIT FEATURE END ---
                else:
                    st.info("No entries match the selected filters.")
            else:
                st.info("No form submissions found.")
        except (FileNotFoundError, json.JSONDecodeError) as e:
            st.info(f"No form submissions found. Error: {str(e)}")
    

        
# Initialize session state for service date check
if 'service_date_checked' not in st.session_state:
    st.session_state.service_date_checked = False

# Initialize session state for duplicate service date confirmation
if 'duplicate_service_date_confirmed' not in st.session_state:
    st.session_state.duplicate_service_date_confirmed = False

# Main content area based on navigation
if st.session_state.nav_selection == "Member Login":
    # Clear admin-specific session state if needed
    st.session_state.is_admin = False
    st.session_state.admin_selection = None
    
    # Only show login form if not already verified
    if not st.session_state.member_verified:
        # Center the login form
        col1, col2, col3 = st.columns([1,2,1])
        with col2:
            # Load the dataframe to get TC email and Medicaid IDs
            try:
                df = pd.read_excel('./Master_db.xlsx')
                
                # Get unique TC emails for username dropdown
                tc_emails = df['TC EMAIL'].dropna().unique().tolist()
                
                # Create a selectbox for username outside the form
                username = st.selectbox("Username (TC Email)", tc_emails)
                
                # Filter Medicaid IDs based on selected TC email
                if username:
                    filtered_df = df[df['TC EMAIL'] == username]
                    medicaid_ids = filtered_df['MedicaidID'].dropna().unique().tolist()
                
                # Build display list: "First Last - MedicaidID"
                medicaid_display = [
                    f"{row['FIRST NAME']} {row['LAST NAME']} - {row['MedicaidID']}"
                    for _, row in filtered_df.iterrows()
                ]
                medicaid_id_map = {
                    f"{row['FIRST NAME']} {row['LAST NAME']} - {row['MedicaidID']}": row['MedicaidID']
                    for _, row in filtered_df.iterrows()
                }
                # Now create the login form with the filtered Medicaid IDs
                with st.form("member_login"):
                    selected_display = st.selectbox("Select Member", medicaid_display if username else [])
                    selected_medicaid_id = medicaid_id_map.get(selected_display, None)
                    # Get the TC name for password verification
                    if username and selected_medicaid_id is not None:
                        tc_row = filtered_df[filtered_df['MedicaidID'] == selected_medicaid_id].iloc[0]
                        expected_password = tc_row.get('TRANSITION COORDINATOR', '')
                    else:
                        expected_password = ""
                    password = st.text_input("Password", type="password")
                    submit = st.form_submit_button("Login")
                    if submit:
                        try:
                            # Validate Medicaid ID format
                            is_valid, error_msg = validate_medicaid_id(selected_medicaid_id)
                            if not is_valid:
                                st.error(error_msg)
                            else:
                                # Check if the password matches the uniform TC password
                                if password == UNIFORM_TC_PASSWORD:
                                    # Get member details from the dataframe
                                    member_details = {
                                        'medicaid_id': selected_medicaid_id,
                                        'member_name': f"{tc_row.get('FIRST NAME', '')} {tc_row.get('LAST NAME', '')}",
                                        'member_id': tc_row.get('MEMBER ID', ''),
                                        'member_dob': tc_row.get('DOB', ''),
                                        'tc_name': tc_row.get('TRANSITION COORDINATOR', ''),
                                        'tc_email': tc_row.get('TC EMAIL', '')
                                    }
                                    st.session_state.member_verified = True
                                    st.session_state.member_data = member_details
                                    # Initialize form section if not already set
                                    if 'current_section' not in st.session_state:
                                        st.session_state.current_section = 0
                                    # Success message
                                    st.success("Login successful!")
                                    # Reset service date check flag
                                    st.session_state.service_date_checked = False
                                    st.session_state.duplicate_service_date_confirmed = False
                                    # Set default tab to Home
                                    st.session_state.member_tab = "Home"
                                    st.rerun()
                                else:
                                    st.error("Invalid password. Please try again.")
                        except Exception as e:
                            st.error(f"Error during login: {str(e)}")
                            print(f"Error during login: {str(e)}")  # Debug log
            except Exception as e:
                st.error(f"Error loading member data: {str(e)}")
                print(f"Error loading member data: {str(e)}")  # Debug log
                
                # Fallback to original login form if data loading fails
                with st.form("member_login_fallback"):
                    username = st.text_input("Username (Medicaid ID)")
                    password = st.text_input("Password", type="password")
                    submit = st.form_submit_button("Login")
                    
                    if submit:
                        try:
                            # Validate Medicaid ID format
                            is_valid, error_msg = validate_medicaid_id(username)
                            if not is_valid:
                                st.error(error_msg)
                            else:
                                # Check if the Medicaid ID exists in the database
                                member_details = get_member_details(username)
                                
                                if not member_details:
                                    st.error("Invalid Medicaid ID. Please try again.")
                                else:
                                    # For demo purposes, accept any password
                                    st.session_state.member_verified = True
                                    st.session_state.member_data = member_details
                                    
                                    # Initialize form section if not already set
                                    if 'current_section' not in st.session_state:
                                        st.session_state.current_section = 0
                                    
                                    # Success message
                                    st.success("Login successful!")
                                    
                                    # Reset service date check flag
                                    st.session_state.service_date_checked = False
                                    st.session_state.duplicate_service_date_confirmed = False
                                    
                                    # Set default tab to Home
                                    st.session_state.member_tab = "Home"
                                    st.rerun()
                        except Exception as e:
                            st.error(f"Error during login: {str(e)}")
                            print(f"Error during login: {str(e)}")  # Debug log
    else:
        # If already verified, show the member dashboard with tabs
        # Initialize member tab if not set
        if 'member_tab' not in st.session_state:
            st.session_state.member_tab = "Home"
        
        # Add logout button at the top
        col1, col2, col3 = st.columns([3, 1, 1])
        with col1:
            # Welcome message
            member_name = st.session_state.member_data.get('member_name', 'Member')
            st.markdown(f"### Welcome, {member_name}! 👋")
        # Remove the Switch Tab button
        with col3:
            if st.button("Logout", use_container_width=True, type="secondary"):
                # Reset all member-related session stateswh
                st.session_state.member_verified = False
                st.session_state.member_data = {}
                st.session_state.form_data = {}
                st.session_state.current_section = 0
                st.session_state.service_date_checked = False
                st.session_state.duplicate_service_date_confirmed = False
                st.session_state.member_tab = "Home"
                st.rerun()
        
        # Create tabs for member dashboard
        tab1, tab2 = st.tabs(["🏠 Home", "📝 New Form"])

        # --- Reset form state if switching to New Form tab ---
        if 'last_member_tab' not in st.session_state:
            st.session_state.last_member_tab = st.session_state.member_tab
        if st.session_state.member_tab != st.session_state.last_member_tab:
            if st.session_state.member_tab == "New Form":
                st.session_state.current_section = 0
                st.session_state.form_data = {}
                st.session_state.service_date_checked = False
                st.session_state.duplicate_service_date_confirmed = False
                st.session_state.selected_service_date = None
            st.session_state.last_member_tab = st.session_state.member_tab
        # ------------------------------------------------------
        
        with tab1:
            st.markdown('<h2 class="subheader">My Submitted Forms</h2>', unsafe_allow_html=True)
            
            # Load log entries
            try:
                with open('log_entries.json', 'r') as f:
                    log_entries = json.load(f)
                
                # Filter entries for current member
                current_medicaid_id = st.session_state.member_data.get('medicaid_id', '')
                member_entries = [
                    entry for entry in log_entries 
                    if (entry.get('medicaid_id') == current_medicaid_id or 
                        entry.get('MEDICAID ID') == current_medicaid_id)
                ]
                
                # Use entries as they are without patching times
                # Remove automatic time patching - let users manually enter times
                patched_entries = member_entries.copy()
                
                if patched_entries:
                    # Dynamically get all columns from all entries
                    all_columns = set()
                    for entry in patched_entries:
                        all_columns.update(entry.keys())
                    all_columns = sorted(list(all_columns))
                    
                    # Define the desired column order
                    desired_order = [
                        "Start time", "Completion time", "Email", "Name", "Is this a new note or amendment?", "REASON FOR AMENDMENT",
                        "DATE OF SERVICE", "Did you travel to/for client?", "TOTAL CLIENT TRAVEL TIME", "CLIENT TRAVEL DETAILS",
                        "TOTAL TRAVEL TIME", "OUTLINE EACH DEST", "ADDITIONAL COMMENTS", "TCM HOURS-ENTER", "TCM UNITS/Minutes",
                        "ICD 10", "CPT CODE", "TRANSITION COORDINATION TASK COMPLETED", "NEXT STEPS/PLAN FOR", "TYPE OF CONTACT",
                        "MEDICAID ID", "MEMBER NAME", "MEMBER ID", "MEMBER DOB"
                    ]
                    # Add any extra columns at the end
                    all_columns = desired_order + [col for col in all_columns if col not in desired_order]
                    # Create DataFrame with ordered columns
                    df = pd.DataFrame(patched_entries, columns=all_columns)
                    st.dataframe(df, use_container_width=True)
                    
                    # Add view details functionality
                    st.markdown("### View Form Details")
                    selected_entry_idx = st.selectbox(
                        "Select a form to view details",
                        options=range(len(patched_entries)),
                        format_func=lambda i: entry_label(patched_entries[i])
                    )
                    
                    if st.button("View Details"):
                        selected_entry = patched_entries[selected_entry_idx]
                        pdf_bytes = dict_to_pdf(selected_entry, title="Form Details")
                        b64_pdf = base64.b64encode(pdf_bytes).decode('utf-8')
                        pdf_display = f'<iframe src="data:application/pdf;base64,{b64_pdf}" width="700" height="900" type="application/pdf"></iframe>'
                        st.markdown(pdf_display, unsafe_allow_html=True)
                        st.download_button(
                            label="Download PDF",
                            data=pdf_bytes,
                            file_name="form_details.pdf",
                            mime="application/pdf"
                        )
                else:
                    st.info("No forms submitted yet.")
                    
                    # # Add a button to create a new form
                    # col1, col2, col3 = st.columns([1, 2, 1])
                    # with col2:
                    #     if st.button("Create Your First Form", type="primary", use_container_width=True):
                    #         st.session_state.member_tab = "New Form"
                    #         st.rerun()
            except (FileNotFoundError, json.JSONDecodeError) as e:
                st.info(f"No forms found. Error: {str(e)}")
                
                # # Add a button to create a new form
                # col1, col2, col3 = st.columns([1, 2, 1])
                # with col2:
                #     if st.button("Create Your First Form", type="primary", use_container_width=True):
                #         st.session_state.member_tab = "New Form"
                #         st.rerun()
        
        with tab2:
            st.markdown('<h2 class="subheader">Create New Form</h2>', unsafe_allow_html=True)
            
            # Check for existing records with the same service date
            if not st.session_state.service_date_checked:
                with st.form("service_date_form"):
                    service_date = st.date_input("Please enter the service date for this form", format="MM/DD/YYYY")
                    submit_date = st.form_submit_button("Continue")
                    
                    if submit_date:
                        # Check if there are existing records with the same service date and medicaid ID
                        try:
                            with open('log_entries.json', 'r') as f:
                                all_entries = json.load(f)
                            
                            # Convert service_date to MM/DD/YYYY string for comparison
                            service_date_str = service_date.strftime("%m/%d/%Y")
                            medicaid_id = st.session_state.member_data.get('medicaid_id', '')
                            
                            # Find entries with matching service date and medicaid ID
                            matching_entries = [
                                entry for entry in all_entries 
                                if (entry.get('DATE OF SERVICE') == service_date_str or 
                                    entry.get('service_date') == service_date_str) and 
                                   (entry.get('MEDICAID ID') == medicaid_id or
                                    entry.get('medicaid_id') == medicaid_id)
                            ]
                            
                            if matching_entries:
                                # Store the service date in session state and mark as checked but not confirmed
                                st.session_state.selected_service_date = service_date
                                st.session_state.service_date_checked = True
                                st.session_state.duplicate_service_date_confirmed = False  # Explicitly set to False
                                st.rerun()  # Rerun to show confirmation dialog
                            else:
                                # No matching entries, proceed with form
                                st.session_state.service_date_checked = True
                                st.session_state.duplicate_service_date_confirmed = True
                                # Store the service date for later use
                                st.session_state.selected_service_date = service_date
                                st.rerun()
                        except (FileNotFoundError, json.JSONDecodeError):
                            # No existing entries file or invalid JSON
                            st.session_state.service_date_checked = True
                            st.session_state.duplicate_service_date_confirmed = True
                            # Store the service date for later use
                            st.session_state.selected_service_date = service_date
                            st.rerun()
            
            # Show confirmation dialog if duplicate service date found and not yet confirmed
            elif st.session_state.service_date_checked and not st.session_state.duplicate_service_date_confirmed:
                # Add a more prominent warning message
                st.markdown("### ⚠️ Duplicate Service Date Detected")
                st.warning(f"**There are already entries for {st.session_state.selected_service_date.strftime('%m/%d/%Y')} for this member.**")
                st.info("Please confirm if you want to continue with this date or choose a different one.")
                
                col1, col2 = st.columns(2)
                with col1:
                    if st.button("✅ Yes, continue with this date", type="primary", use_container_width=True):
                        st.session_state.duplicate_service_date_confirmed = True
                        st.rerun()
                with col2:
                    if st.button("🔄 No, choose a different date", type="secondary", use_container_width=True):
                        st.session_state.service_date_checked = False
                        st.session_state.duplicate_service_date_confirmed = False
                        st.rerun()
            
            # Only show progress bar for non-Administrative notes
            elif st.session_state.duplicate_service_date_confirmed:
                # Create the progress steps
                def create_progress_bar(current_section, total_sections=TOTAL_SECTIONS):
                    html = '<div class="step-container">'
                    for i in range(total_sections):  # Changed from 1 to 0-based indexing
                        if i < current_section:
                            html += f'<div class="step completed">{i}</div>'
                        elif i == current_section:
                            html += f'<div class="step active">{i}</div>'
                        else:
                            html += f'<div class="step">{i}</div>'
                        
                        if i < total_sections - 1:  # Changed condition for last step
                            if i < current_section:
                                html += '<div class="step-line completed"></div>'
                            else:
                                html += '<div class="step-line"></div>'
                    html += '</div>'
                    return html

                # Only show progress bar for non-Administrative notes
                if st.session_state.get('form_data', {}).get('note_category') != "Administrative":
                    # Progress bar with custom styling
                    st.markdown(create_progress_bar(st.session_state.current_section), unsafe_allow_html=True)
                    # st.markdown(f'<p class="progress-text">Section {st.session_state.current_section} of {TOTAL_SECTIONS}</p>', unsafe_allow_html=True)
                    
                    # Add Previous/Next navigation
                    prev_next_cols = st.columns([1, 8, 1]) # Adjust column ratios
                    with prev_next_cols[0]:
                        st.markdown('<div class="nav-arrow">', unsafe_allow_html=True)
                        if st.button("←", 
                                    disabled=st.session_state.current_section == 0,
                                    use_container_width=True,
                                    key="prev_button"):
                            st.session_state.current_section = max(0, st.session_state.current_section - 1)
                            st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)
                    
                    with prev_next_cols[2]:
                        st.markdown('<div class="nav-arrow">', unsafe_allow_html=True)
                        if st.button("→", 
                                    disabled=st.session_state.current_section == TOTAL_SECTIONS - 1,
                                    use_container_width=True,
                                    key="next_button"):
                            st.session_state.current_section = min(TOTAL_SECTIONS - 1, st.session_state.current_section + 1)
                            st.rerun()
                        st.markdown('</div>', unsafe_allow_html=True)
                
                # Before the form, check if we need to set a note category
                if 'note_category' not in st.session_state:
                    st.session_state.note_category = "Billable- TCM"  # Default value
                
                medicaid_id = st.session_state.member_data.get('medicaid_id', '')
                        
                # Get member details if available
                member_details = st.session_state.member_data

                # Section 0 (Demographics) - NEW SECTION
                if st.session_state.current_section == 0:
                    # st.markdown('<h2 class="subheader">DEMOGRAPHICS</h2>', unsafe_allow_html=True)
                    
                    with st.form("demographics_form"):
                        # Row 1: Medicaid ID and Member Name
                        col1, col2 = st.columns(2)
                        with col1:
                            st.text_input(
                                "MEDICAID ID *",
                                value=medicaid_id,
                                disabled=True,
                                key="medicaid_id_display"
                            )
                        with col2:
                            st.text_input(
                                "MEMBER NAME *",
                                value=member_details.get('member_name', ''),
                                disabled=True,
                                key="member_name_display"
                            )
                        
                        # Row 2: Member DOB and Member ID
                        col1, col2 = st.columns(2)
                        with col1:
                            # Handle DOB properly
                            try:
                                if 'member_dob' in member_details and member_details['member_dob']:
                                    if isinstance(member_details['member_dob'], (str, datetime, date)):
                                        dob_value = pd.to_datetime(member_details['member_dob']).date()
                                    else:
                                        dob_str = str(member_details['member_dob'])
                                        if dob_str and dob_str.lower() != 'nan':
                                            dob_value = pd.to_datetime(dob_str).date()
                                        else:
                                            dob_value = datetime.now().date()
                                else:
                                    dob_value = datetime.now().date()
                                
                                st.date_input(
                                    "MEMBER DOB *",
                                    value=dob_value,
                                    disabled=True,
                                    key="member_dob_display",
                                    format="MM/DD/YYYY"
                                )
                            except Exception as e:
                                st.error(f"Error processing date of birth: {str(e)}")
                                # Fallback to current date
                                st.date_input(
                                    "MEMBER DOB *",
                                    value=datetime.now().date(),
                                    disabled=True,
                                    key="member_dob_display_fallback",
                                    format="MM/DD/YYYY"
                                )
                        with col2:
                            st.text_input(
                                "MEMBER ID", 
                                value=str(int(float(member_details.get('member_id', 0)))),  # Convert to int to remove decimal points
                                disabled=True,
                                key="member_id_display", 
                                help="Member ID must be a numerical value"
                            )
                        
                        # Row 3: Transition Coordinator Name and Email
                        col1, col2 = st.columns(2)
                        with col1:
                            st.text_input("Transition Coordinator Name", value=member_details.get('tc_name', ''), disabled=True)
                        with col2:
                            st.text_input("Transition Coordinator Email", value=member_details.get('tc_email', ''), disabled=True)
                        
                        # Row 4: Service Date
                        st.date_input(
                            "DATE OF SERVICE", 
                            value=st.session_state.selected_service_date,
                            disabled=True,
                            format="MM/DD/YYYY"
                        )
                        # REMOVED START TIME FROM SECTION 0
                        # Add Next button
                        submitted = st.form_submit_button("Next")
                        if submitted:
                            # Move to next section
                            st.session_state.current_section = 1
                            st.rerun()

                # Only show note type selection in section 1
                elif st.session_state.current_section == 1:
                    # Create a container outside the form to handle note type selection
                    note_type_container = st.container()
                    with note_type_container:
                        
                        # Default values
                        start_time_val = st.session_state.get('form_data', {}).get('start_time', dt_time(9, 0))
                        end_time_val = st.session_state.get('form_data', {}).get('end_time', dt_time(17, 0))

                        # Ensure form_data exists in session_state
                        if 'form_data' not in st.session_state:
                            st.session_state.form_data = {}

                        # Time input widgets
                        
                        temp_note_category = st.radio(
                            "**[1] Type of Note**", 
                            ["Administrative", "Billable- TCM"],
                            index=0 if st.session_state.note_category == "Administrative" else 1,
                            key="note_category_selector",  # Use a different key
                            horizontal=True  # Display options horizontally
                        )
                        # Update session state when selection changes
                        if temp_note_category != st.session_state.note_category:
                            st.session_state.note_category = temp_note_category
                            # Reset form data when changing note type
                            if 'form_data' in st.session_state:
                                st.session_state.form_data = {}
                            st.rerun()
                    # Inside the forms
                    if st.session_state.note_category == "Administrative":
                        # Administrative section start_time
                        start_time_val = parse_time(st.session_state.get('form_data', {}).get('start_time', None), default=dt_time(9, 0))
                        start_time_input = st.time_input(
                            "START TIME (EST)",
                            value=start_time_val,
                            key="start_time_editable_section1"
                        )
                        st.session_state.form_data['start_time'] = start_time_input

                        

                        note_type = st.radio(
                            "[2] Is this a new note or an amendment to correct a previous note?",
                            ["New Note", "Amendment"],
                            key="admin_note_type"
                        )
                        travel_to_client = st.radio("[3] DID YOU TRAVEL TO/FOR CLIENT", ["Yes", "No"], index=1, key="admin_travel_radio_2", horizontal=True)
                        with st.form("admin_form"):
                            amendment_reason = ""
                            if note_type == "Amendment":
                                amendment_reason = st.text_area(
                                    "REASON FOR FORM AMENDMENT",
                                    height=100,
                                    key="amendment_reason_admin"
                                )
                            is_disabled = (travel_to_client == "No")
                            travel_time_val = 0.0 if is_disabled else st.session_state.get('admin_travel_time', 0.0)
                            travel_time = st.number_input(
                                "[3.1] TOTAL CLIENT TRAVEL TIME (15 MIN INCREMENTS)",
                                min_value=0.0,
                                max_value=24.0,
                                step=0.25,
                                value=travel_time_val,
                                disabled=is_disabled,
                                key="admin_travel_time"
                            )
                            travel_details = ""
                            if travel_to_client == "Yes":
                                st.markdown("""
                                [3.2] In this section, please specify the details of all your travel destinations, 
                                including the starting and ending addresses for each stop.
                                """)
                                travel_details = st.text_area("Outline each destination to and from locations")
                            st.markdown("**[4.1] ADMINISTRATIVE TYPE**")
                            admin_type = st.radio(
                                "Select Administrative Type",
                                options=["MEETING", "Training", "Travel"],
                                key="admin_type_radio",
                                label_visibility="collapsed"
                            )
                            st.markdown("**[4.2] PLEASE ENTER ADMINISTRATIVE WORK COMPLETED**")
                            admin_comments = st.text_area(
                                "Enter administrative work details",
                                height=200,
                                help="Provide details about the administrative work completed",
                                key="admin_comments_direct"
                            )

                            # Administrative section end_time
                            end_time_val = parse_time(st.session_state.get('form_data', {}).get('end_time', None), default=dt_time(17, 0))
                            end_time_input = st.time_input(
                                "END TIME (EST)",
                                value=end_time_val,
                                key="end_time_editable_section1"
                            )
                            st.session_state.form_data['end_time'] = end_time_input
                            total_travel_time_hidden = 0.0
                            travel_locations_hidden = ""
                            travel_comments_hidden = ""
                            tasks_completed_hidden = "Administrative task"
                            next_steps_hidden = "N/A for Administrative note"
                            contact_types_hidden = ["DOCUMENTATION"]
                            admin_submitted = st.form_submit_button("Submit")
                            if admin_submitted:
                                # Validate that required times are entered
                                start_time_entered = st.session_state.form_data.get('start_time')
                                end_time_entered = st.session_state.form_data.get('end_time')
                                
                                if not start_time_entered:
                                    st.error("Please enter a Start Time before submitting the form.")
                                    st.stop()
                                
                                if not end_time_entered:
                                    st.error("Please enter an End Time before submitting the form.")
                                    st.stop()
                                
                                # Process Administrative form submission
                                # REMOVE ANY VALIDATION CODE FOR MEMBER ID
                                
                                # Save the form data
                                form_data = {
                                    'medicaid_id': medicaid_id,
                                    'member_name': member_details.get('member_name', ''),
                                    'member_id': member_details.get('member_id', 0),
                                    'member_dob': member_details.get('member_dob', ''),
                                    'note_type': note_type,
                                    'service_date': st.session_state.selected_service_date,
                                    'travel_to_client': travel_to_client,
                                    'note_category': st.session_state.note_category,
                                    'tc_name': member_details.get('tc_name', ''),
                                    'tc_email': member_details.get('tc_email', '')
                                }
                                
                                # Add user-entered times from the form
                                if 'start_time' in st.session_state.form_data:
                                    start_time = st.session_state.form_data['start_time']
                                    if hasattr(start_time, 'strftime'):
                                        form_data['start_time'] = start_time.strftime("%H:%M")
                                    else:
                                        form_data['start_time'] = str(start_time)
                                
                                if 'end_time' in st.session_state.form_data:
                                    end_time = st.session_state.form_data['end_time']
                                    if hasattr(end_time, 'strftime'):
                                        form_data['end_time'] = end_time.strftime("%H:%M")
                                    else:
                                        form_data['end_time'] = str(end_time)
                                
                                # Add amendment reason if applicable
                                if note_type == "Amendment":
                                    form_data['amendment_reason'] = amendment_reason
                                
                                # Add travel details if applicable
                                if travel_to_client == "Yes":
                                    form_data['travel_time'] = travel_time
                                    form_data['travel_details'] = travel_details
                                else:
                                    form_data['travel_time'] = 0.0
                                    form_data['travel_details'] = ""
                                
                                # Add admin type if applicable
                                form_data['admin_type'] = admin_type
                                form_data['admin_comments'] = admin_comments
                                
                                # Add default values for required fields in other sections
                                form_data['total_travel_time'] = total_travel_time_hidden
                                form_data['travel_locations'] = travel_locations_hidden
                                form_data['travel_comments'] = travel_comments_hidden
                                form_data['tasks_completed'] = tasks_completed_hidden
                                form_data['next_steps'] = next_steps_hidden
                                form_data['contact_types'] = contact_types_hidden
                                
                                # For Administrative notes, create the complete entry and submit
                                # Use the new mapping function to convert form data to required fields
                                entry = map_form_data_to_required_fields(form_data)
                                # entry["timestamp"] = datetime.now().strftime("%m/%d/%Y %H:%M:%S")
                                
                                # Add the extra fields that are not in REQUIRED_FIELDS but needed for the database
                                extra_fields = [
                                    'medicaid_id', 'member_name', 'member_dob', 'member_id',
                                    'tc_name', 'tc_email', 'service_date', 'start_time', 'end_time'
                                ]
                                for field in extra_fields:
                                    if field in form_data:
                                        entry[field] = form_data[field]
                                
                                st.session_state.log_entries.append(entry)
                                save_entries()
                                st.session_state.form_data = {}
                                st.success("Administrative form submitted successfully!")
                                
                                # Reset form-related session states but keep user logged in
                                st.session_state.current_section = 0
                                st.session_state.service_date_checked = False
                                st.session_state.duplicate_service_date_confirmed = False
                                st.session_state.form_data = {}
                                
                                # Switch to Home tab to show the submitted form
                                st.session_state.member_tab = "Home"
                                st.rerun()
                    else:  # Billable- TCM form
                        existing_data = st.session_state.get('form_data', {})
                        # if 'start_time' not in st.session_state.form_data:
                        #     st.session_state.form_data['start_time'] = datetime.now(pytz.timezone('US/Eastern')).time()
                        
                        # Get medicaid_id from session state
                        start_time_input = st.time_input(
                                "START TIME (EST)",
                                value=start_time_val,
                                key="start_time_editable_section1"
                            )
                        
                        # Store the time input value in session state after the widget
                        st.session_state.form_data['start_time'] = start_time_input
                        
                        
                        note_type = st.radio(
                            "**[2] Is this a new note or an amendment to correct a previous note?**",
                            ["New Note", "Amendment"],
                            index=0 if existing_data.get('note_type') != "Amendment" else 1,
                            key="tcm_note_type"
                        )
                        travel_to_client = st.radio("**[3] DID YOU TRAVEL TO/FOR CLIENT**", ["Yes", "No"], index=1, key="tcm_travel_radio_2", horizontal=True)
                                    
                        # TCM-specific fields with auto-adjustment logic
                        st.markdown("**[4] TCM HOURS AND UNITS**")

                        col1, col2 = st.columns(2)

                        with col1:
                            tcm_hours_input = st.number_input(
                                "TCM HOURS (15 min increments)",
                                min_value=0.0,
                                max_value=24.0,
                                step=0.25,
                                value=existing_data.get('tcm_hours_input', 0.0),
                                key="tcm_hours_input",
                                help="Enter time in 15-minute increments (e.g., 1.25 = 1 hour 15 minutes)"
                            )

                        # Auto-adjustment logic: Adjusted Hours = Entered Hours + 0.25
                        if tcm_hours_input > 0:
                            adjusted_hours = tcm_hours_input + 0.25
                            auto_units = int(adjusted_hours * 4)  # Units = Adjusted Hours * 4
                            
                            # Display the adjusted values
                            st.info(f"📊 **Auto-Adjustment Applied:**\n"
                                    f"- Entered Hours: {tcm_hours_input}\n"
                                    f"- Adjusted Hours: {adjusted_hours}\n" 
                                    f"- Calculated Units: {auto_units}")
                            
                            # Set the adjusted values
                            tcm_hours = adjusted_hours
                            tcm_units = auto_units
                        else:
                            tcm_hours = tcm_hours_input
                            tcm_units = 0

                        with col2:
                            # Display the auto-calculated units (read-only or editable)
                            tcm_units_display = st.number_input(
                                "TCM UNITS (Auto-calculated)",
                                min_value=0,
                                max_value=96,
                                step=1,
                                value=tcm_units,
                                key="tcm_units_display",
                                help="Automatically calculated: (Hours + 0.25) × 4",
                                disabled=True  # Make read-only since it's auto-calculated
                            )

                        # Manual override option
                        # manual_override = st.checkbox("Override auto-calculation", key="manual_override_tcm")

                        # if manual_override:
                        #     tcm_units_manual = st.number_input(
                        #         "MANUAL TCM UNITS",
                        #         min_value=0,
                        #         max_value=96,
                        #         step=1,
                        #         value=tcm_units,
                        #         key="tcm_units_manual",
                        #         help="Enter custom units to override auto-calculation"
                        #     )
                        #     # Replace the auto-calculated value with manual input
                        #     tcm_units = tcm_units_manual

                        # Final values to use in your form submission
                        final_tcm_hours = tcm_hours_input  # Store original entered hours
                        final_tcm_units = tcm_units        # Store final units (auto-calc or manual)
                        #############################################################################################

                        with st.form("tcm_form_section1"):
                            amendment_reason = ""
                            if note_type == "Amendment":
                                st.markdown('<p class="section-number">1.1a)</p>', unsafe_allow_html=True)
                                amendment_reason = st.text_area(
                                    "REASON FOR FORM AMENDMENT",
                                    value=existing_data.get('amendment_reason', ''),
                                    height=100,
                                    key="amendment_reason"
                                )
                            
                            
                            
                            is_disabled = (travel_to_client == "No")
                            
                            if is_disabled:
                                travel_time_val = 0.0
                            else:
                                travel_time_val = st.session_state.get('tcm_travel_time', existing_data.get('travel_time', 0.0))

                            travel_time = st.number_input(
                                "[3.1] TOTAL CLIENT TRAVEL TIME (15 MIN INCREMENTS)",
                                min_value=0.0,
                                max_value=24.0,
                                step=0.25,
                                value=float(travel_time_val),
                                disabled=is_disabled,
                                key="tcm_travel_time"
                            )

                            # if travel_to_client == "Yes":
                            st.markdown("""
                            [3.2] In this section, please specify the details of all your travel destinations, 
                            including the starting and ending addresses for each stop.
                            """)
                            travel_details = st.text_area("Outline each destination to and from locations", value=existing_data.get('travel_details', ''))
                        
                            cpt_options = ["Please select",
                                "T1017 TRANSITION COORDINATION",
                                    "T2038 HOUSEHOLD SET UP TIME",
                                    "Administrative"]
                            cpt_index = 0
                            if existing_data.get('cpt_code') in cpt_options:
                                cpt_index = cpt_options.index(existing_data.get('cpt_code'))
                            
                            cpt_code = st.selectbox(
                                "**[5] CPT CODE**",
                                cpt_options,
                                index=cpt_index
                            )
                        
                            icd_10 = st.checkbox("**ICD 10 - R99**", 
                            value=existing_data.get('icd_10', False),
                            help="International Classification of Diseases, 10th revision code")

                            # Add a Next button at the bottom of the form
                            tcm_submitted = st.form_submit_button("Next")
                            
                            if tcm_submitted:
                                # Validate that required times are entered
                                start_time_entered = st.session_state.form_data.get('start_time')
                                
                                if not start_time_entered:
                                    st.error("Please enter a Start Time before proceeding.")
                                    st.stop()
                                
                                
                                # Process TCM form section 1 submission
                                # REMOVE ANY VALIDATION CODE FOR MEMBER ID
                                # end_time = datetime.now(pytz.timezone('US/Eastern')).time()
                                
                                # Save the form data
                                form_data = {
                                    'medicaid_id': medicaid_id,
                                    'member_name': member_details.get('member_name', ''),
                                    'member_id': member_details.get('member_id', 0),
                                    'member_dob': member_details.get('member_dob', ''),
                                    'note_type': note_type,
                                    'service_date': st.session_state.selected_service_date,
                                    'travel_to_client': travel_to_client,
                                    'note_category': st.session_state.note_category,
                                    'tc_name': member_details.get('tc_name', ''),
                                    'tc_email': member_details.get('tc_email', '')
                                }
                                
                                # Add user-entered times from the form
                                if 'start_time' in st.session_state.form_data:
                                    start_time = st.session_state.form_data['start_time']
                                    if hasattr(start_time, 'strftime'):
                                        form_data['start_time'] = start_time.strftime("%H:%M")
                                    else:
                                        form_data['start_time'] = str(start_time)
                                
                                if 'end_time' in st.session_state.form_data:
                                    end_time = st.session_state.form_data['end_time']
                                    if hasattr(end_time, 'strftime'):
                                        form_data['end_time'] = end_time.strftime("%H:%M")
                                    else:
                                        form_data['end_time'] = str(end_time)
                                
                                # Add amendment reason if applicable
                                if note_type == "Amendment" and 'amendment_reason' in locals():
                                    form_data['amendment_reason'] = amendment_reason
                                
                                # Add travel details if applicable
                                if travel_to_client == "Yes":
                                    form_data['travel_time'] = travel_time
                                    form_data['travel_details'] = travel_details
                                else:
                                    form_data['travel_time'] = 0.0
                                    form_data['travel_details'] = ""
                                
                                # Add TCM details
                                form_data['tcm_hours'] = tcm_hours
                                form_data['tcm_units'] = tcm_units
                                form_data['icd_10'] = icd_10
                                form_data['cpt_code'] = cpt_code
                                
                                # Update session state
                                st.session_state.form_data.update(form_data)  # Use update to preserve existing data
                                
                                if tcm_submitted:
                                    # Move to next section for sequential navigation
                                    st.session_state.current_section = 2
                                    st.rerun()


elif st.session_state.nav_selection == "Support":
    # st.markdown('<h1 class="main-title">Support</h1>', unsafe_allow_html=True)
    
    # Support page content
    st.markdown("""
    ### Contact Information
    - **Email:** sreevani.patil@focuscares.com
    - **Phone:** 1-XXX-XXX-XXXX
    
    ### Hours of Operation
    Monday - Friday: 9:00 AM - 5:00 PM MST
    
    
    ### Submit a Support Ticket
    """)
    
    # Support ticket form
    with st.form("support_ticket"):
        issue_type = st.selectbox(
            "Issue Type",
            ["Technical Problem", "Feature Request", "Data Issue", "Other"]
        )
        description = st.text_area("Description", height=25)
        priority = st.select_slider(
            "Priority",
            options=["Low", "Medium", "High", "Urgent"]
        )
        submitted = st.form_submit_button("Submit Ticket")
        
        if submitted:
            st.success("Support ticket submitted successfully!")

elif st.session_state.nav_selection == "Admin":
    # Check if user is admin
    if not st.session_state.is_admin:
        # Show admin login form
        # st.markdown('<h1 class="main-title">Admin Login</h1>', unsafe_allow_html=True)
        
        # Center the login form
        col1, col2, col3 = st.columns([1,2,1])
        with col2:
            with st.form("admin_login"):
                username = st.text_input("Username")
                password = st.text_input("Password", type="password")
                submit = st.form_submit_button("Login")
                
                if submit:
                    # Check credentials
                    if username in ADMIN_CREDENTIALS:
                        hashed_password = hashlib.sha256(password.encode()).hexdigest()
                        if hashed_password == ADMIN_CREDENTIALS[username]:
                            st.session_state.is_admin = True
                            st.session_state.admin_selection = "View Forms"  # Changed from "View Submitted Forms"
                            st.success("Admin login successful!")
                            st.rerun()
                        else:
                            st.error("Invalid password")
                    else:
                        st.error("Invalid username")
    else:
        
        # If no specific admin section is selected, show dashboard
        if not st.session_state.get('admin_selection'):
            # Create three columns for metrics
            col1, col2, col3 = st.columns(3)
            
            # Load log entries to calculate metrics
            try:
                with open('log_entries.json', 'r') as f:
                    log_entries = json.load(f)
                    
                # Calculate metrics
                total_forms = len(log_entries)
                
                # Count forms submitted today
                today = datetime.now().strftime("%m/%d/%Y")
                forms_today = sum(1 for entry in log_entries if entry.get('timestamp', '').startswith(today))
                
                # Count unique users (medicaid IDs)
                unique_users = len(set(entry.get('medicaid_id') for entry in log_entries if 'medicaid_id' in entry))
                
                # Display metrics
                # with col1:
                #     st.metric("Total Forms", total_forms)
                
                # with col2:
                #     st.metric("Forms Today", forms_today)
                
                # with col3:
                #     st.metric("Active Users", unique_users)
                    
            except (FileNotFoundError, json.JSONDecodeError):
                # If no entries file exists or is invalid
                with col1:
                    st.metric("Total Forms", 0)
                
                with col2:
                    st.metric("Forms Today", 0)
                
                with col3:
                    st.metric("Active Users", 0)
            
            # Remove the duplicate navigation buttons here


# After the service date confirmation check, add this debug information
if st.session_state.duplicate_service_date_confirmed:
    # Debug information
    # st.write(f"Current section: {st.session_state.current_section}")
    
    # Make sure form_data exists
    if 'form_data' not in st.session_state:
        st.session_state.form_data = {}
    
    # Section 1 (already implemented)
    if st.session_state.current_section == 1:
        # Your existing section 1 code
        pass
    
    
    # Section 3
    elif st.session_state.current_section == 2:
        with st.form("tasks_form"):
            st.markdown('<h2 class="subheader">[6] TASKS COMPLETED</h2>', unsafe_allow_html=True)
            tasks_completed_text = st.text_area(
                "**[6.1 & 6.2] ENTER TASKS COMPLETED**",
                value="",
                height=50,
                help="Describe all transition coordination tasks completed during this session"
            )
            next_steps = st.text_area(
                "**[6.3] ENTER NEXT STEPS AND FOLLOW-UP PLAN**",
                height=25,
                help="Detail the planned next steps and follow-up actions"
            )
            st.markdown("**[7] TYPE OF CONTACT**")
            contact_types = st.multiselect(
                "SELECT TYPE(S) OF CONTACT",
                options=[
                    "CALL",
                    "EMAIL",
                    "IN PERSON",
                    "DOCUMENTATION",
                    "VIRTUAL",
                    "Other"
                ]
            )
            other_contact_type = None
            if "Other" in contact_types:
                other_contact_type = st.text_input(
                    "Please specify other contact type(s)"
                )
            submitted = st.form_submit_button("Next")
            if submitted:
                section_data = {
                    'tasks_completed': tasks_completed_text,
                    'next_steps': next_steps,
                    'contact_types': contact_types
                }
                if other_contact_type:
                    section_data['other_contact_type'] = other_contact_type
                st.session_state.form_data.update(section_data)
                st.session_state.current_section += 1
                st.rerun()

    # Section 4 (First Contact)
    elif st.session_state.current_section == 3:
        st.markdown('<h2 class="subheader">[8] FIRST CONTACT</h2>', unsafe_allow_html=True)
        
        with st.form("first_contact_form"):  # Wrap in form
            first_contact_name = st.text_input("**[8.1] FULL NAME**")
            
            first_contact_email = st.text_input("**[8.2] EMAIL**")
            
            first_contact_phone = st.text_input(
                "**[8.3] PHONE NUMBER**",
                help="Format: +1 XXX-XXX-XXXX (must include +1)",
                placeholder="+1 XXX-XXX-XXXX"
            )
            
            first_contact_outcome = st.radio(
                "**[8.4] OUTCOME**",
                options=[
                    "DISCONNECTED/WRONG NUMBER",
                    "EMAIL",
                    "LEFT MESSAGE",
                    "NO ANSWER",
                    "SPOKE TO CONTACT",
                    "VOICEMAIL FULL",
                    "Other"
                ]
            )
            
            first_contact_other_outcome = None
            if first_contact_outcome == "Other":
                first_contact_other_outcome = st.text_input("Please specify other outcome")
            
            need_second_contact = st.radio(
                "**[8.5] DO YOU HAVE ANOTHER CONTACT TO ENTER?**",
                ["Yes", "No"]
            )
            
            # Add the Next button inside the form
            submitted = st.form_submit_button("Next")
            if submitted:
                # Save the contact information to session state
                contact_data = {
                    'contact_name': first_contact_name,
                    'contact_email': first_contact_email,
                    'contact_phone': first_contact_phone,
                    'contact_outcome': first_contact_outcome
                }
                if first_contact_outcome == "Other" and first_contact_other_outcome:
                    contact_data['other_outcome'] = first_contact_other_outcome
                
                st.session_state.form_data.update({'first_contact': contact_data})
                
                # Determine next section based on need_second_contact
                if need_second_contact == "Yes":
                    st.session_state.current_section = 4  # Go to second contact
                else:
                    st.session_state.current_section = 7  # Skip to final section
                st.rerun()

    # Section 5 (Second Contact)
    elif st.session_state.current_section == 4:
        st.markdown('<h2 class="subheader">[9] SECOND CONTACT</h2>', unsafe_allow_html=True)
        
        with st.form("second_contact_form_section5"):
            second_contact_name = st.text_input("**[9.1] FULL NAME**", key="second_contact_name_sec5")
            
            second_contact_email = st.text_input("**[9.2] EMAIL**", key="second_contact_email_sec5")
            
            second_contact_phone = st.text_input(
                "**[9.3] PHONE NUMBER**",
                help="Format: +1 XXX-XXX-XXXX (must include +1)",
                placeholder="+1 XXX-XXX-XXXX",
                key="second_contact_phone_sec5"
            )
            
            second_contact_outcome = st.radio(
                "**[9.4] OUTCOME**",
                options=[
                    "SPOKE TO CONTACT",
                    "LEFT MESSAGE",
                    "DISCONNECTED/WRONG NUMBER",
                    "NO ANSWER",
                    "VOICEMAIL FULL",
                    "Other"
                ],
                key="second_contact_outcome_sec5"
            )
            
            second_contact_other_outcome = None
            if second_contact_outcome == "Other":
                second_contact_other_outcome = st.text_input(
                    "Please specify other outcome (Second Contact)",
                    key="second_contact_other_outcome_sec5"
                )
            
            need_third_contact = st.radio(
                "**[9.5] DO YOU NEED TO ENTER ANOTHER CONTACT?**",
                ["Yes", "No"],
                key="need_third_contact_sec5"
            )
            
            submitted = st.form_submit_button("Next")
            if submitted:
                # Save contact information
                contact_data = {
                    'contact_name': second_contact_name,
                    'contact_email': second_contact_email,
                    'contact_phone': second_contact_phone,
                    'contact_outcome': second_contact_outcome
                }
                if second_contact_outcome == "Other" and second_contact_other_outcome:
                    contact_data['other_outcome'] = second_contact_other_outcome
                
                st.session_state.form_data.update({'second_contact': contact_data})
                
                if submitted:
                    # Navigate based on need_third_contact
                    if need_third_contact == "Yes":
                        st.session_state.current_section = 5
                    else:
                        st.session_state.current_section = 8
                    st.rerun()

    # Section 6 (Third Contact)
    elif st.session_state.current_section == 5:
        st.markdown('<h2 class="subheader">[10] THIRD CONTACT</h2>', unsafe_allow_html=True)
        
        with st.form("third_contact_form_section6"):
            third_contact_name = st.text_input("**[10.1] FULL NAME**", key="third_contact_name_sec6")
            
            third_contact_email = st.text_input("**[10.2] EMAIL**", key="third_contact_email_sec6")
            
            third_contact_phone = st.text_input(
                "**[10.3] PHONE NUMBER**",
                help="Format: +1 XXX-XXX-XXXX (must include +1)",
                placeholder="+1 XXX-XXX-XXXX",
                key="third_contact_phone_sec6"
            )
            
            third_contact_outcome = st.radio(
                "**[10.4] OUTCOME**",
                options=[
                    "SPOKE TO CONTACT",
                    "LEFT MESSAGE",
                    "DISCONNECTED/WRONG NUMBER",
                    "NO ANSWER",
                    "VOICEMAIL FULL",
                    "Other"
                ],
                key="third_contact_outcome_sec6"
            )
            
            third_contact_other_outcome = None
            if third_contact_outcome == "Other":
                third_contact_other_outcome = st.text_input(
                    "Please specify other outcome (Third Contact)",
                    key="third_contact_other_outcome_sec6"
                )
            
            need_fourth_contact = st.radio(
                "**[10.5] DO YOU NEED TO ENTER ANOTHER CONTACT?**",
                ["Yes", "No"],
                key="need_fourth_contact_sec6"
            )
            
            submitted = st.form_submit_button("Next")
            if submitted:
                # Save contact information
                contact_data = {
                    'contact_name': third_contact_name,
                    'contact_email': third_contact_email,
                    'contact_phone': third_contact_phone,
                    'contact_outcome': third_contact_outcome
                }
                if third_contact_outcome == "Other" and third_contact_other_outcome:
                    contact_data['other_outcome'] = third_contact_other_outcome
                
                st.session_state.form_data.update({'third_contact': contact_data})
                
                if submitted:
                    # Navigate based on need_fourth_contact
                    if need_fourth_contact == "Yes":
                        st.session_state.current_section = 6
                    else:
                        st.session_state.current_section = 8
                    st.rerun()

    # Section 7 (Fourth Contact)
    elif st.session_state.current_section == 6:
        st.markdown('<h2 class="subheader">[11] FOURTH CONTACT</h2>', unsafe_allow_html=True)
        
        with st.form("fourth_contact_form_section7"):
            fourth_contact_name = st.text_input("**[11.1] FULL NAME**", key="fourth_contact_name_sec7")
            
            fourth_contact_email = st.text_input("**[11.2] EMAIL**", key="fourth_contact_email_sec7")
            
            fourth_contact_phone = st.text_input(
                "**[11.3] PHONE NUMBER**",
                help="Format: +1 XXX-XXX-XXXX (must include +1)",
                placeholder="+1 XXX-XXX-XXXX",
                key="fourth_contact_phone_sec7"
            )
            
            fourth_contact_outcome = st.radio(
                "**[11.4] OUTCOME**",
                options=[
                    "SPOKE TO CONTACT",
                    "LEFT MESSAGE",
                    "DISCONNECTED/WRONG NUMBER",
                    "NO ANSWER",
                    "VOICEMAIL FULL",
                    "Other"
                ],
                key="fourth_contact_outcome_sec7"
            )
            
            fourth_contact_other_outcome = None
            if fourth_contact_outcome == "Other":
                fourth_contact_other_outcome = st.text_input(
                    "**[11.5] PLEASE SPECIFY OTHER OUTCOME (FOURTH CONTACT)**",
                    key="fourth_contact_other_outcome_sec7"
                )
            
            submitted = st.form_submit_button("Next")
            if submitted:
                # Save contact information
                contact_data = {
                    'contact_name': fourth_contact_name,
                    'contact_email': fourth_contact_email,
                    'contact_phone': fourth_contact_phone,
                    'contact_outcome': fourth_contact_outcome
                }
                if fourth_contact_outcome == "Other" and fourth_contact_other_outcome:
                    contact_data['other_outcome'] = fourth_contact_other_outcome
                
                st.session_state.form_data.update({'fourth_contact': contact_data})
                
                # Move to final section
                st.session_state.current_section = 7
                st.rerun()

    # Section 8 (Final Section)
    elif st.session_state.current_section == 7:
        st.markdown('<h2 class="subheader">ADMINISTRATIVE COMMENTS</h2>', unsafe_allow_html=True)
        
        with st.form("final_form_section8"):
            st.markdown("**PLEASE ENTER ADMINISTRATIVE WORK COMPLETED**")
            admin_comments = st.text_area(
                "Enter administrative work details",
                height=25,
                help="Provide details about the administrative work completed",
                key="admin_comments_sec8"
            )

            # Final section end_time
            end_time_val = parse_time(st.session_state.get('form_data', {}).get('end_time', None), default=dt_time(17, 0))
            end_time_input = st.time_input(
                "END TIME (EST)",
                value=end_time_val,
                key="end_time_editable_section1"
            )
            st.session_state.form_data['end_time'] = end_time_input
            submitted = st.form_submit_button("Submit")
            
            if submitted:
                # Validate that required times are entered
                start_time_entered = st.session_state.form_data.get('start_time')
                end_time_entered = st.session_state.form_data.get('end_time')
                
                if not start_time_entered:
                    st.error("Please enter a Start Time before submitting the form.")
                    st.stop()
                
                if not end_time_entered:
                    st.error("Please enter an End Time before submitting the form.")
                    st.stop()
                
                # Save the admin comments
                st.session_state.form_data.update({"admin_comments": admin_comments if admin_comments else ""})
                
                # Use only user-entered times, no automatic system time
                # Ensure the form data contains the actual user-entered times
                if 'start_time' in st.session_state.form_data:
                    # Convert time object to string format if needed
                    start_time = st.session_state.form_data['start_time']
                    if hasattr(start_time, 'strftime'):
                        st.session_state.form_data['start_time'] = start_time.strftime("%H:%M")
                
                if 'end_time' in st.session_state.form_data:
                    # Convert time object to string format if needed
                    end_time = st.session_state.form_data['end_time']
                    if hasattr(end_time, 'strftime'):
                        st.session_state.form_data['end_time'] = end_time.strftime("%H:%M")
                
                # Build entry with all required fields using the new mapping function
                entry = map_form_data_to_required_fields(st.session_state.form_data)
                # Remove 'Id' and 'timestamp' if present
                entry.pop('Id', None)
                entry.pop('timestamp', None)
                
                # Only use user-entered times, no automatic fallback
                # If times are missing, they will remain as entered (or empty)
                
                st.session_state.log_entries.append(entry)
                save_entries()
                st.session_state.form_data = {}
                st.success("Form submitted successfully!")
                
                # Reset form-related session states but keep user logged in
                st.session_state.current_section = 0
                st.session_state.service_date_checked = False
                st.session_state.duplicate_service_date_confirmed = False
                
                # Switch to Home tab to show the submitted form
                st.session_state.member_tab = "Home"
                st.rerun()
                # For save_continue, just save the data and stay on current section


